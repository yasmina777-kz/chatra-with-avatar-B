import openpyxl

def parse_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    all_sheets = {}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = []

        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                row_data = [str(cell) if cell is not None else "" for cell in row]
                rows.append(" | ".join(row_data))

        all_sheets[sheet_name] = "\n".join(rows)

    return {
        "type": "excel",
        "sheets": all_sheets
    }